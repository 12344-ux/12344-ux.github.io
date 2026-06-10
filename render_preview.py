#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Genera una vista previa HTML fiel del Excel generado (merges, bordes, anchos)."""
import html
import openpyxl
from openpyxl.utils import get_column_letter

SRC = 'RECONOCIMIENTO DE APRENDIZAJES PREVIOS - DILIGENCIADO.xlsx'
OUT = 'vista_previa.html'
MAXROW = 712

wb = openpyxl.load_workbook(SRC)
ws = wb.active

# mapa de celdas cubiertas por merges y su ancla
covered = {}
anchor = {}
for m in ws.merged_cells.ranges:
    for r in range(m.min_row, m.max_row + 1):
        for c in range(m.min_col, m.max_col + 1):
            if (r, c) != (m.min_row, m.min_col):
                covered[(r, c)] = True
    anchor[(m.min_row, m.min_col)] = (m.max_row - m.min_row + 1,
                                      m.max_col - m.min_col + 1)

NCOLS = 8
# anchos de columna -> px
widths = []
for c in range(1, NCOLS + 1):
    dim = ws.column_dimensions[get_column_letter(c)]
    w = dim.width or 10
    widths.append(int(w * 7))

out = []
out.append('<!doctype html><html><head><meta charset="utf-8">')
out.append('<title>Vista previa - Diagnóstico de Aprendizajes Previos</title>')
out.append('''<style>
body{background:#e9ebee;font-family:Calibri,Arial,sans-serif;padding:20px;}
table{border-collapse:collapse;background:#fff;margin:0 auto;box-shadow:0 2px 10px rgba(0,0,0,.15);}
td{border:1px solid #000;padding:3px 5px;vertical-align:top;font-size:11px;
   word-wrap:break-word;overflow:hidden;}
.hd{background:#d8d8d8;font-weight:bold;text-align:center;vertical-align:middle;}
.comp{font-weight:bold;text-align:center;vertical-align:middle;color:#1f3864;font-size:12px;white-space:pre-line;}
.title{font-weight:bold;text-align:center;font-size:15px;}
.info{font-weight:bold;}
.mark{text-align:center;vertical-align:middle;}
</style></head><body>''')

# colgroup
out.append('<table>')
out.append('<colgroup>')
for w in widths:
    out.append(f'<col style="width:{w}px">')
out.append('</colgroup>')

for r in range(1, MAXROW + 1):
    out.append('<tr>')
    for c in range(1, NCOLS + 1):
        if (r, c) in covered:
            continue
        cell = ws.cell(row=r, column=c)
        val = cell.value
        text = '' if val is None else html.escape(str(val))
        span = anchor.get((r, c))
        attrs = ''
        if span:
            rs, cs = span
            if rs > 1:
                attrs += f' rowspan="{rs}"'
            if cs > 1:
                attrs += f' colspan="{cs}"'
        cls = ''
        if r in (16, 17):
            cls = 'hd'
        elif r == 2:
            cls = 'title'
        elif c == 1 and r >= 18 and span and span[0] > 1:
            cls = 'comp'
        elif c in (6, 7) and r >= 18:
            cls = 'mark'
        clsattr = f' class="{cls}"' if cls else ''
        out.append(f'<td{attrs}{clsattr}>{text}</td>')
    out.append('</tr>')
out.append('</table></body></html>')

open(OUT, 'w').write('\n'.join(out))
print('Generado', OUT)
