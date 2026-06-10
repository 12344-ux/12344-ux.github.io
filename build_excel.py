#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Rellena la plantilla del Diagnostico de Aprendizajes Previos con los datos
extraidos del PDF (competencias.json), manteniendo estructura y estilo.
"""
import json
import math
import copy
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

SRC = 'RECONOCIMIENTO DE APRENDIZAJES PREVIOS V2 (1).xlsx'
OUT = 'RECONOCIMIENTO DE APRENDIZAJES PREVIOS - DILIGENCIADO.xlsx'

competencias = json.load(open('competencias.json'))

wb = openpyxl.load_workbook(SRC)
ws = wb.active

# ---------- estilos ----------
thin = Side(style='thin', color='FF000000')
medium = Side(style='medium', color='FF000000')
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

def border_block(left_medium=False, right_medium=False):
    return Border(left=medium if left_medium else thin,
                  right=medium if right_medium else thin,
                  top=thin, bottom=thin)

font_comp = Font(name='Calibri', size=10, bold=True, color='FF000000')
font_code = Font(name='Calibri', size=10, bold=True, color='FF1F3864')
font_item = Font(name='Calibri', size=8, color='FF000000')
font_mark = Font(name='Calibri', size=11, bold=True, color='FF000000')

al_comp = Alignment(horizontal='center', vertical='center', wrap_text=True)
al_item = Alignment(horizontal='left', vertical='top', wrap_text=True)
al_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

# ancho aprox de caracteres por columna (fuente 8) para estimar alto de fila
CHARS = {'B': 44, 'C': 43, 'D': 41, 'E': 53}
COLS_ITEM = ['B', 'C', 'D', 'E']


def est_lines(text, width_chars):
    if not text:
        return 1
    total = 0
    for seg in str(text).split('\n'):
        seg = seg if seg else ' '
        total += max(1, math.ceil(len(seg) / width_chars))
    return total


# ---------- limpiar zona de datos (fila 18 en adelante) ----------
# quitar merges de la zona de datos/firmas, conservar cabecera (<=17)
for rng in list(ws.merged_cells.ranges):
    if rng.min_row >= 18:
        ws.unmerge_cells(str(rng))

# borrar contenido y estilo de la zona de datos
max_clear = max(ws.max_row, 1100)
for row in ws.iter_rows(min_row=18, max_row=max_clear, min_col=1, max_col=28):
    for c in row:
        c.value = None
        c.border = Border()
        c.fill = PatternFill()
        if ws.row_dimensions[c.row].height:
            ws.row_dimensions[c.row].height = None

# ---------- escribir bloques ----------
def style_data_row(r):
    """Aplica borde a A..H del registro y devuelve nada."""
    for ci, col in enumerate('ABCDEFGH', start=1):
        cell = ws.cell(row=r, column=ci)
        lm = (col == 'A')
        rm = (col == 'H')
        cell.border = border_block(left_medium=lm, right_medium=rm)


row = 18
for comp in competencias:
    ra = comp['resultados_aprendizaje'] or ['']
    proc = comp['conocimientos_proceso'] or ['']
    saber = comp['conocimientos_saber'] or ['']
    crit = comp['criterios_evaluacion'] or ['']
    h = max(len(ra), len(proc), len(saber), len(crit), 1)
    r0, r1 = row, row + h - 1

    # estilo + bordes de todas las filas del bloque
    for rr in range(r0, r1 + 1):
        style_data_row(rr)

    # --- columna A: competencia (merge) ---
    ws.merge_cells(start_row=r0, start_column=1, end_row=r1, end_column=1)
    a = ws.cell(row=r0, column=1)
    dur = f" ({comp['duracion']})" if comp['duracion'] else ''
    a.value = f"{comp['codigo']}  (Versión {comp['version']}){dur}\n\n{comp['denominacion']}"
    a.font = font_comp
    a.alignment = al_comp

    # --- columnas B,C,D,E: un item por fila ---
    for col_idx, (col_letter, items) in enumerate(
            [('B', ra), ('C', proc), ('D', saber), ('E', crit)], start=2):
        for i, val in enumerate(items):
            cell = ws.cell(row=r0 + i, column=col_idx)
            cell.value = val
            cell.font = font_item
            cell.alignment = al_item

    # --- F (SI), G (NO), H (OBSERVACIONES): merge por competencia, vacios ---
    for col_idx in (6, 7, 8):
        ws.merge_cells(start_row=r0, start_column=col_idx, end_row=r1, end_column=col_idx)
        cell = ws.cell(row=r0, column=col_idx)
        cell.font = font_mark
        cell.alignment = al_center

    # --- alto de filas estimado ---
    for i in range(h):
        rr = r0 + i
        lines = 1
        for col in COLS_ITEM:
            items = {'B': ra, 'C': proc, 'D': saber, 'E': crit}[col]
            if i < len(items):
                lines = max(lines, est_lines(items[i], CHARS[col]))
        ws.row_dimensions[rr].height = max(15.0, lines * 11.5 + 3)

    row = r1 + 1

# ---------- firmas ----------
row += 1
firmas = [
    'NOMBRE Y FIRMA DEL APRENDIZ:',
    'NOMBRE Y FIRMA DEL INSTRUCTOR:',
]
for txt in firmas:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    c = ws.cell(row=row, column=1)
    c.value = txt
    c.font = Font(name='Calibri', size=11, bold=True)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[row].height = 34
    row += 2

wb.save(OUT)
print('Guardado:', OUT)
print('Ultima fila de datos aprox:', row)
print('Competencias escritas:', len(competencias))
